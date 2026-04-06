import datetime
import json
import os
import re
import unicodedata

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "ASISTENCIA", "PERSONAL OBRERO 2026.xlsx")
OUTPUT_FILES = [
    os.path.join(BASE_DIR, "asistencia_data.js"),
    os.path.join(BASE_DIR, "src", "asistencia_data.js"),
    os.path.join(BASE_DIR, "docs", "asistencia_data.js"),
]

EXPECTED_HEADERS = [
    "N",
    "DNI N",
    "APELLIDO PATERNO",
    "APELLIDO MATERNO",
    "NOMBRES",
    "CATEGORIA",
    "OCUPACION",
    "FECHA DE INGRESO",
]


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[^A-Za-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def to_clean_string(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def format_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime.date):
        return value.strftime("%d/%m/%Y")
    return to_clean_string(value)


def build_person_id(item, dni, row_number):
    if dni:
        return f"DNI-{dni}"
    if item:
        return f"ITEM-{item}"
    return f"ROW-{row_number}"


def load_personal_from_excel(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active

    headers = [normalize_text(sheet.cell(3, col).value) for col in range(1, 9)]
    expected = [normalize_text(header) for header in EXPECTED_HEADERS]

    if headers != expected:
        print("Advertencia: los encabezados no coinciden exactamente con lo esperado.")
        print("Leido   :", headers)
        print("Esperado:", expected)

    personal = []
    seen_ids = set()

    for row in range(4, sheet.max_row + 1):
        values = [sheet.cell(row, col).value for col in range(1, 9)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        item = to_clean_string(values[0])
        dni = to_clean_string(values[1])
        ape_pat = to_clean_string(values[2])
        ape_mat = to_clean_string(values[3])
        nombres = to_clean_string(values[4])
        categoria = to_clean_string(values[5])
        ocupacion = to_clean_string(values[6])
        fecha_ingreso = format_date(values[7])

        person_id = build_person_id(item, dni, row)
        if person_id in seen_ids:
            person_id = f"{person_id}-{row}"
        seen_ids.add(person_id)

        nombre_completo = " ".join(part for part in [ape_pat, ape_mat, nombres] if part).strip()

        personal.append(
            {
                "id": person_id,
                "n": item,
                "dni": dni,
                "apellidoPaterno": ape_pat,
                "apellidoMaterno": ape_mat,
                "nombres": nombres,
                "categoria": categoria,
                "ocupacion": ocupacion,
                "fechaIngreso": fecha_ingreso,
                "nombreCompleto": nombre_completo,
            }
        )

    return personal


def write_js_data(personal, out_file):
    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    payload = json.dumps(personal, ensure_ascii=False)
    content = (
        f"// Auto-generado desde ASISTENCIA/PERSONAL OBRERO 2026.xlsx ({len(personal)} registros)\n"
        f"const PERSONAL_ASISTENCIA = {payload};\n"
        "if (typeof window !== 'undefined') { window.PERSONAL_ASISTENCIA = PERSONAL_ASISTENCIA; }\n"
    )
    with open(out_file, "w", encoding="utf-8") as f:
        f.write(content)


if __name__ == "__main__":
    if not os.path.exists(EXCEL_PATH):
        raise SystemExit(f"No se encontro el Excel: {EXCEL_PATH}")

    personal_data = load_personal_from_excel(EXCEL_PATH)
    for out in OUTPUT_FILES:
        write_js_data(personal_data, out)
        print(f"OK: {out}")

    print(f"Total personal cargado: {len(personal_data)}")

