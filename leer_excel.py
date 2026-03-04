import openpyxl, json

wb = openpyxl.load_workbook(r'C:\Users\Jose Alonso\IntelliJ_IDEA\work_manage_app\Metrado - Estructuras.xlsx', data_only=True)
ws = wb.active

partidas = []
for i in range(1, ws.max_row + 1):
    col_a = ws.cell(row=i, column=1).value
    col_b = ws.cell(row=i, column=2).value
    col_d = ws.cell(row=i, column=4).value
    if col_a and str(col_a).strip().lower() == 'partida':
        codigo = str(col_b).strip() if col_b else ''
        nombre = str(col_d).strip() if col_d else ''
        if codigo and nombre:
            partidas.append({'codigo': codigo, 'nombre': nombre})

def get_categoria_y_color(nombre):
    n = ' '.join(nombre.upper().split())  # normaliza espacios
    if 'ZAPATA' in n or 'SUB ZAPATA' in n:
        return 'ZAPATAS', '#E74C3C'
    if 'CIMIENTO' in n:
        return 'CIMIENTOS', '#E67E22'
    if 'SOLADO' in n or 'FALSO PISO' in n:
        return 'SOLADOS', '#F39C12'
    if 'COLUMNETA' in n:
        return 'COLUMNETAS', '#9B59B6'
    if 'COLUMNA' in n:
        return 'COLUMNAS', '#8E44AD'
    if 'VIGUETA' in n:
        return 'VIGUETAS', '#3498DB'
    if 'VIGA' in n:
        return 'VIGAS', '#2980B9'
    if 'LOSA ALIGERADA' in n or 'LOSAS ALIGERADAS' in n or 'LADRILLO HUECO' in n:
        return 'LOSAS_ALIGERADAS', '#27AE60'
    if 'LOSA MACIZA' in n or 'LOSAS MACIZAS' in n:
        return 'LOSAS_MACIZAS', '#2ECC71'
    if 'ESCALERA' in n:
        return 'ESCALERAS', '#16A085'
    if 'PLACA' in n:
        return 'PLACAS', '#1ABC9C'
    if 'MURO' in n:
        return 'MUROS', '#34495E'
    if 'PARAPETO' in n:
        return 'PARAPETOS', '#95A5A6'
    if 'MESON' in n or 'MESONES' in n:
        return 'MESONES', '#BDC3C7'
    if 'SOBRECIMIENTO' in n:
        return 'SOBRECIMIENTOS', '#D35400'
    if 'EXCAVAC' in n or 'RELLENO' in n or 'ACARREO' in n or 'ELIMINACION' in n or 'NIVELACION' in n:
        return 'MOVIMIENTO_TIERRAS', '#795548'
    if 'JUNTA' in n or 'IMPERMEABIL' in n:
        return 'JUNTAS', '#607D8B'
    if 'CURADO' in n:
        return 'CURADO', '#00BCD4'
    return 'OTROS', '#9E9E9E'

for p in partidas:
    cat, color = get_categoria_y_color(p['nombre'])
    p['categoria'] = cat
    p['color'] = color

out_path = r'C:\Users\Jose Alonso\IntelliJ_IDEA\work_manage_app\src\partidas.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(partidas, f, ensure_ascii=False, indent=2)

print(f"OK: {len(partidas)} partidas guardadas en {out_path}")
for p in partidas:
    print(f"  [{p['categoria']:<20}] {p['codigo']:<35} {p['nombre'][:60]}")

