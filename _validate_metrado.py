import openpyxl
import re

wb = openpyxl.load_workbook('DATA PARTIDAS/Metrado - Estructuras.xlsx', data_only=True)
ws = wb.active

partidas = []
for row in range(13, ws.max_row+1):
    a_val = ws.cell(row=row, column=1).value
    b_val = ws.cell(row=row, column=2).value
    d_val = ws.cell(row=row, column=4).value
    m_val = ws.cell(row=row, column=13).value
    if a_val == 'Partida' and b_val and d_val:
        partidas.append((row, b_val, d_val, m_val))

ok_count = 0
fail_count = 0
for idx, (row, num, desc, metrado_raw) in enumerate(partidas):
    if metrado_raw:
        m_str = str(metrado_raw).replace(',', '')
        m = re.match(r'([\d]+\.?\d*)', m_str)
        metrado_val = float(m.group(1)) if m else None
    else:
        metrado_val = None

    if idx + 1 < len(partidas):
        end_row = partidas[idx+1][0]
    else:
        end_row = ws.max_row + 1

    header_row = row + 2
    parcial_col = 12  # column L

    # Find where 'Resumen' starts (if any)
    resumen_row = end_row
    for r in range(header_row + 2, end_row):
        i_val = ws.cell(row=r, column=9).value
        if i_val and str(i_val).strip() == 'Resumen':
            resumen_row = r
            break

    total = 0.0
    data_start = header_row + 2
    for r in range(data_start, resumen_row):
        v = ws.cell(row=r, column=parcial_col).value
        if v is not None:
            try:
                total += float(v)
            except:
                pass

    if metrado_val is not None:
        diff = abs(total - metrado_val)
        if diff < 0.5:
            ok_count += 1
        else:
            fail_count += 1
            print(f"DIFF: {num} | Suma={total:.2f} vs Met={metrado_val:.2f} (diff={diff:.2f})")

print(f"\nResultado: OK={ok_count}, DIFF={fail_count}, Total={len(partidas)}")

