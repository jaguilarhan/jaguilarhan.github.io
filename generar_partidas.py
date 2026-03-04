"""
generar_partidas.py - Lee PRESUPUESTO TOTAL y genera partidas.json para la app.
Solo incluye partidas con metrado (las reales, no títulos).
Clasifica automáticamente por categoría basándose en el nombre.
"""
import openpyxl, json, os, re

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, "DATA PARTIDAS", "PRESUPUESTO TOTAL - 01 INFRAESTRUCTURA.xlsx")
OUT = os.path.join(BASE, "src", "partidas.json")

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb.active

# ── Colores por categoría ────────────────────────────────────────
COLORES = {
    "ZAPATAS":          "#E74C3C",
    "CIMIENTOS":        "#E67E22",
    "SOBRECIMIENTOS":   "#F39C12",
    "COLUMNAS":         "#8E44AD",
    "COLUMNETAS":       "#9B59B6",
    "PLACAS":           "#1ABC9C",
    "VIGAS":            "#2980B9",
    "VIGUETAS":         "#3498DB",
    "LOSAS_ALIGERADAS": "#27AE60",
    "LOSAS_MACIZAS":    "#2ECC71",
    "ESCALERAS":        "#16A085",
    "MUROS":            "#607D8B",
    "MUROS_SOST":       "#455A64",
    "PARAPETOS":        "#795548",
    "MESONES":          "#8D6E63",
    "MOVIMIENTO_TIERRAS":"#A1887F",
    "CONCRETO_SIMPLE":  "#FF7043",
    "ARQUITECTURA":     "#42A5F5",
    "INST_SANITARIAS":  "#26C6DA",
    "INST_ELECTRICAS":  "#FFA726",
    "INST_COMUNICACION":"#AB47BC",
    "SEGURIDAD":        "#EF5350",
    "AMBIENTAL":        "#66BB6A",
    "OTROS":            "#90A4AE",
}

def clasificar(codigo, nombre):
    """Clasifica partida por categoría basándose en código y nombre."""
    n = nombre.upper()
    c = codigo

    # Por nombre del elemento estructural
    if "ZAPATA" in n:                    return "ZAPATAS"
    if "CIMIENTO" in n and "SOBRE" not in n: return "CIMIENTOS"
    if "SOBRECIMIENTO" in n:             return "SOBRECIMIENTOS"
    if "COLUMNETA" in n:                 return "COLUMNETAS"
    if "COLUMNA" in n:                   return "COLUMNAS"
    if "PLACA" in n:                     return "PLACAS"
    if "VIGUETA" in n:                   return "VIGUETAS"
    if "VIGA" in n:                      return "VIGAS"
    if "LOSA" in n and "ALIGER" in n:    return "LOSAS_ALIGERADAS"
    if "LOSA" in n and "MACIZ" in n:     return "LOSAS_MACIZAS"
    if "LOSA" in n:                      return "LOSAS_ALIGERADAS"
    if "ESCALERA" in n:                  return "ESCALERAS"
    if "MURO" in n and "SOST" in n:      return "MUROS_SOST"
    if "MURO" in n or "TABIQUE" in n:    return "MUROS"
    if "PARAPETO" in n:                  return "PARAPETOS"
    if "MESON" in n or "MESÓN" in n:     return "MESONES"

    # Por código de especialidad (primeros 2 dígitos)
    esp = c.split(".")[0] if "." in c else c[:2]
    if esp == "01":
        # Dentro de estructuras, subcategorizar
        if "EXCAVAC" in n or "RELLENO" in n or "NIVELAC" in n or "ACARREO" in n or "ELIMINAC" in n:
            return "MOVIMIENTO_TIERRAS"
        if "SOLADO" in n or "FALSO PISO" in n or "SUB ZAPATA" in n:
            return "CONCRETO_SIMPLE"
        return "OTROS"
    if esp == "02": return "OTROS"           # Obras provisionales
    if esp == "03": return "ARQUITECTURA"
    if esp == "04": return "INST_SANITARIAS"
    if esp == "05": return "INST_ELECTRICAS"
    if esp == "06": return "INST_COMUNICACION"
    if esp == "07": return "AMBIENTAL"
    if esp == "08": return "SEGURIDAD"
    return "OTROS"

# ── Primero recolectamos todos los items para construir jerarquía ──
items_raw = []
for row in range(13, ws.max_row + 1):
    codigo_raw = ws.cell(row, 1).value
    nombre_raw = ws.cell(row, 2).value
    metrado    = ws.cell(row, 5).value

    if not codigo_raw or not nombre_raw:
        continue

    codigo = str(codigo_raw).strip()
    nombre = str(nombre_raw).strip()
    if not codigo or not nombre or nombre == 'None':
        continue

    items_raw.append({
        "codigo": codigo,
        "nombre": nombre,
        "metrado": metrado,
        "fila": row,
    })

# ── Construir dict de padres para la ruta jerárquica ──
nombres_por_codigo = {}
for item in items_raw:
    nombres_por_codigo[item["codigo"]] = item["nombre"]

def obtener_especialidad(codigo):
    """Obtiene el código de especialidad (2 dígitos) y su nombre."""
    parts = codigo.split(".")
    esp_cod = parts[0]
    return esp_cod, nombres_por_codigo.get(esp_cod, "")

def obtener_ruta(codigo):
    """Obtiene la ruta jerárquica completa."""
    parts = codigo.split(".")
    ruta = []
    for i in range(1, len(parts)):
        parent_cod = ".".join(parts[:i])
        if parent_cod in nombres_por_codigo:
            ruta.append(nombres_por_codigo[parent_cod])
    return ruta

# ── Generar partidas.json (solo las que tienen metrado) ──
partidas = []
for item in items_raw:
    metrado = item["metrado"]
    if metrado is None or metrado == '' or metrado == 0:
        continue

    codigo = item["codigo"]
    nombre = item["nombre"]
    cat = clasificar(codigo, nombre)
    esp_cod, esp_nombre = obtener_especialidad(codigo)
    ruta = obtener_ruta(codigo)

    partidas.append({
        "codigo": codigo,
        "nombre": nombre,
        "categoria": cat,
        "color": COLORES.get(cat, "#90A4AE"),
        "metrado": float(metrado) if metrado else 0,
        "especialidad": esp_nombre,
        "especialidad_cod": esp_cod,
        "ruta": ruta,
    })

# Ordenar por código
partidas.sort(key=lambda p: [int(x) for x in p["codigo"].split(".")])

# ── Leer PARTIDAS NUEVAS del Excel adicional ────────────────────
EXCEL_NUEVAS = os.path.join(BASE, "DATA PARTIDAS", "PRESUPUESTO - PARTIDAS NUEVAS -W2026-v2.xlsx")
if os.path.exists(EXCEL_NUEVAS):
    print("Leyendo partidas nuevas...")
    wb2 = openpyxl.load_workbook(EXCEL_NUEVAS, data_only=True)
    ws2 = wb2['JUSTIFICACION']

    # Hoja JUSTIFICACION: A=Item, B=Descripcion, E=Unidad, F=Metrado
    # Títulos no tienen metrado, partidas reales sí
    items_nuevas = []
    nombres_nuevas = {}
    for row in range(13, ws2.max_row + 1):
        cod_raw = ws2.cell(row, 1).value
        nom_raw = ws2.cell(row, 2).value
        metrado = ws2.cell(row, 6).value  # columna F = metrado

        if not cod_raw or not nom_raw:
            continue
        codigo = str(cod_raw).strip()
        nombre = str(nom_raw).strip()
        if not codigo or not nombre or nombre == 'None':
            continue

        nombres_nuevas[codigo] = nombre
        items_nuevas.append({
            "codigo": codigo,
            "nombre": nombre,
            "metrado": metrado,
            "fila": row,
        })

    def obtener_ruta_nuevas(codigo):
        parts = codigo.split(".")
        ruta = []
        for i in range(1, len(parts)):
            parent_cod = ".".join(parts[:i])
            if parent_cod in nombres_nuevas:
                ruta.append(nombres_nuevas[parent_cod])
        return ruta

    count_nuevas = 0
    for item in items_nuevas:
        metrado = item["metrado"]
        if metrado is None or metrado == '' or metrado == 0:
            continue

        codigo = item["codigo"]
        nombre = item["nombre"]
        cat = clasificar(codigo, nombre)
        ruta = obtener_ruta_nuevas(codigo)
        esp_nombre = ruta[0] if ruta else "PARTIDAS NUEVAS"

        partidas.append({
            "codigo": codigo,
            "nombre": nombre,
            "categoria": cat,
            "color": COLORES.get(cat, "#90A4AE"),
            "metrado": float(metrado) if metrado else 0,
            "especialidad": esp_nombre,
            "especialidad_cod": codigo.split(".")[0],
            "ruta": ruta,
            "nueva": True,
        })
        count_nuevas += 1

    print(f"  Partidas nuevas encontradas: {count_nuevas}")
    wb2.close()
else:
    print(f"No se encontro archivo de partidas nuevas: {EXCEL_NUEVAS}")

# Re-ordenar incluyendo las nuevas
partidas.sort(key=lambda p: [int(x) for x in p["codigo"].split(".")])

# ── Guardar ──
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(partidas, f, ensure_ascii=False, indent=2)

# ── Resumen ──
print(f"Partidas generadas: {len(partidas)}")
print(f"Archivo: {OUT}")
print()

# Conteo por especialidad
from collections import Counter
esp_count = Counter(p["especialidad"] for p in partidas)
print("Por especialidad:")
for esp, cnt in esp_count.most_common():
    print(f"  {cnt:4d}  {esp}")

print()
cat_count = Counter(p["categoria"] for p in partidas)
print("Por categoria:")
for cat, cnt in cat_count.most_common():
    print(f"  {cnt:4d}  {cat}")

# ── Generar partidas_data.js ────────────────────────────────────
OUT_JS = os.path.join(BASE, "src", "partidas_data.js")
with open(OUT_JS, "w", encoding="utf-8") as f:
    f.write(f"// Auto-generado: {len(partidas)} partidas\n")
    f.write("const PARTIDAS = ")
    json.dump(partidas, f, ensure_ascii=False)
    f.write(";\n")
print(f"JS generado: partidas_data.js ({len(partidas)} partidas)")

