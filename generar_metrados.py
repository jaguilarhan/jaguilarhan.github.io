"""
Genera metrados_data.js desde el Excel 'Metrado - Estructuras.xlsx'
Extrae las 69 partidas con sus tablas de detalle (volumétrico o acero)
"""
import openpyxl
import re
import json

wb = openpyxl.load_workbook('DATA PARTIDAS/Metrado - Estructuras.xlsx', data_only=True)
ws = wb.active

# 1. Encontrar todas las partidas
partidas_raw = []
for row in range(13, ws.max_row + 1):
    a_val = ws.cell(row=row, column=1).value
    b_val = ws.cell(row=row, column=2).value
    d_val = ws.cell(row=row, column=4).value
    m_val = ws.cell(row=row, column=13).value
    if a_val == 'Partida' and b_val and d_val:
        partidas_raw.append((row, str(b_val).strip(), str(d_val).strip(), str(m_val).strip() if m_val else ''))

partidas = []
for idx, (row, num, desc, metrado_raw) in enumerate(partidas_raw):
    # Extraer valor numérico y unidad del metrado
    metrado_val = 0.0
    unidad = ''
    if metrado_raw:
        m = re.match(r'([\d,]+\.?\d*)\s*(.*)', metrado_raw.replace(',', ''))
        if m:
            metrado_val = float(m.group(1))
            unidad = m.group(2).strip()

    # Determinar fin de tabla (siguiente partida o fin de archivo)
    if idx + 1 < len(partidas_raw):
        end_row = partidas_raw[idx + 1][0]
    else:
        end_row = ws.max_row + 1

    header_row = row + 2

    # Detectar tipo de tabla por columna H del header
    h_val = ws.cell(row=header_row, column=8).value
    h_str = str(h_val).strip() if h_val else ''

    if 'Var' in h_str:  # N Var. → acero
        tipo = 'acero'
        columnas = ['descripcion', 'diametro', 'n_estr', 'n_elem', 'n_var', 'l_varilla', 'l_total', 'peso_m', 'peso']
        col_map = {
            'descripcion': 2,   # B
            'diametro': 5,      # E
            'n_estr': 6,        # F
            'n_elem': 7,        # G
            'n_var': 8,         # H
            'l_varilla': 9,     # I
            'l_total': 10,      # J
            'peso_m': 11,       # K
            'peso': 12          # L
        }
        parcial_col = 12  # L = Peso
    else:
        tipo = 'volumetrico'
        columnas = ['descripcion', 'n_estr', 'n_elem', 'area', 'largo', 'ancho', 'alto', 'parcial']
        col_map = {
            'descripcion': 2,   # B
            'n_estr': 6,        # F
            'n_elem': 7,        # G
            'area': 8,          # H
            'largo': 9,         # I
            'ancho': 10,        # J
            'alto': 11,         # K
            'parcial': 12       # L
        }
        parcial_col = 12  # L = Parcial

    # Encontrar sección de Resumen (si existe)
    resumen_row = end_row
    for r in range(header_row + 2, end_row):
        i_val = ws.cell(row=r, column=9).value
        if i_val and str(i_val).strip() == 'Resumen':
            resumen_row = r
            break

    # Extraer filas de datos (detalle)
    detalle = []
    data_start = header_row + 2  # skip header + empty row
    for r in range(data_start, resumen_row):
        fila = {}
        tiene_datos = False
        for col_name, col_num in col_map.items():
            v = ws.cell(row=r, column=col_num).value
            if v is not None:
                tiene_datos = True
                if col_name in ('descripcion', 'diametro'):
                    fila[col_name] = str(v).strip()
                else:
                    try:
                        fila[col_name] = round(float(v), 4)
                    except:
                        fila[col_name] = str(v).strip()
            else:
                fila[col_name] = '' if col_name in ('descripcion', 'diametro') else None
        if tiene_datos:
            detalle.append(fila)

    # Extraer resumen (solo para acero)
    resumen = []
    if resumen_row < end_row and tipo == 'acero':
        for r in range(resumen_row + 2, end_row):  # skip "Resumen" header
            i_val = ws.cell(row=r, column=9).value
            j_val = ws.cell(row=r, column=10).value
            k_val = ws.cell(row=r, column=11).value
            l_val = ws.cell(row=r, column=12).value
            if i_val and str(i_val).strip() not in ('', 'ø'):
                try:
                    resumen.append({
                        'diametro': str(i_val).strip(),
                        'l_total': round(float(j_val), 2) if j_val else 0,
                        'peso_m': round(float(k_val), 4) if k_val else 0,
                        'peso': round(float(l_val), 2) if l_val else 0
                    })
                except:
                    pass

    # Validación: suma de parciales
    suma_parcial = sum(
        (f.get('parcial') or f.get('peso') or 0)
        for f in detalle
        if isinstance(f.get('parcial', f.get('peso')), (int, float))
    )

    partida_obj = {
        'id': num.replace('.', '_'),
        'numero': num,
        'descripcion': desc,
        'metrado_total': round(metrado_val, 2),
        'unidad': unidad,
        'tipo': tipo,
        'columnas': columnas,
        'detalle': detalle,
        'suma_verificacion': round(suma_parcial, 2)
    }
    if resumen:
        partida_obj['resumen'] = resumen

    partidas.append(partida_obj)

# Generar archivo JS
js_content = "// Generado automáticamente desde Metrado - Estructuras.xlsx\n"
js_content += "// Total partidas: {}\n".format(len(partidas))
js_content += "const METRADOS_PARTIDAS = " + json.dumps(partidas, ensure_ascii=False, indent=2) + ";\n"

# Guardar en raíz
with open('metrados_data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

# Guardar en docs/
with open('docs/metrados_data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"✅ Generado metrados_data.js con {len(partidas)} partidas")
print(f"   Volumétricas: {sum(1 for p in partidas if p['tipo'] == 'volumetrico')}")
print(f"   Acero: {sum(1 for p in partidas if p['tipo'] == 'acero')}")

# Validación final
ok = 0
for p in partidas:
    diff = abs(p['suma_verificacion'] - p['metrado_total'])
    if diff < 0.5:
        ok += 1
    else:
        print(f"   ⚠️ {p['numero']}: suma={p['suma_verificacion']} vs total={p['metrado_total']}")
print(f"   Validación: {ok}/{len(partidas)} OK")

